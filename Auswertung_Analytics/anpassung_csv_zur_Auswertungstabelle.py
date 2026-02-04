# UTF-8
import pandas as pd
from tqdm import tqdm
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, NamedStyle

def connect(file_list) -> object:
    for i in range(len(file_list)):
        if i == 0:
            df = pd.read_csv(file_list[i], sep=',', header=0, encoding='utf-8')
        else:
            df2 = pd.read_csv(file_list[i], sep=',', header=0, encoding='utf-8')
            df = pd.concat([df, df2], axis=0, ignore_index=True)
    return df

def convert(df, file) -> str:
    df['Knotenpunkt'] = '01'
    col = {'start occurrence date':'DATUM', 
           'Knotenpunkt':'KNOTENPUNKT', 
           'from section':'VON_ARM', 
           'to section':'NACH_ARM', 
           'start occurrence time':'VON_UHR', 
           'end occurrence time':'BIS_UHR'}
    df = df.rename(columns=col)

    kn = df[['DATUM', 'KNOTENPUNKT', 'VON_ARM', 'NACH_ARM', 'VON_UHR', 'BIS_UHR']]
    # kn = kn.groupby(['DATUM', 'VON_ARM', 'NACH_ARM', 'VON_UHR']).first().reset_index()
    kn = kn.drop_duplicates(['DATUM', 'VON_ARM', 'NACH_ARM', 'VON_UHR'])

    kn = kn.join(pd.DataFrame({
        'Pkw':0,
        'K-Rad':0, 
        'Rad':0, 
        'Lfw':0, 
        'Lkw':0, 
        'Lz/Sz':0, 
        'Strab':0, 
        'Bus':0, 
        'Fg':0, 
        'Sonder':0
        }, index=df.index))

    for id, row in tqdm(df.iterrows()):
        # Finde den Index in kn (nach reset_index ist es 0-basiert)
        matching_rows = kn[(kn['VON_UHR'] == row['VON_UHR']) & 
                           (kn['VON_ARM'] == row['VON_ARM']) & 
                           (kn['NACH_ARM'] == row['NACH_ARM'])]
        if not matching_rows.empty:
            idx = matching_rows.index[0]  # Erster Treffer (sollte eindeutig sein nach drop_duplicates)
    # Zuweisen der Messzahlen der Erhebung den Zellen
            typ = row['classification']
            count = row['count']
            if typ == 'bicycle':
                kn.loc[idx, 'Rad'] = count
            elif typ == 'bus':
                kn.loc[idx, 'Bus'] = count
            elif typ == 'car':
                kn.loc[idx, 'Pkw'] = count
            elif typ == 'motorcycle':
                kn.loc[idx, 'K-Rad'] = count
            elif typ == 'person':
                kn.loc[idx, 'Fg'] = count
            elif typ == 'train':
                kn.loc[idx, 'Strab'] = count
            elif typ == 'truck':
                kn.loc[idx, 'Lkw'] = count
    # save als Excel
    file = file.replace('.csv','_bueffeeTable.xlsx', 1)
    with pd.ExcelWriter(file, mode='w', engine='openpyxl') as writer:
        kn.to_excel(writer, sheet_name='DB-01', index=False)
    # fiziere erste Zeile
    workbook = load_workbook(file)
    worksheet = workbook['DB-01']
    worksheet.freeze_panes = "A2"
    workbook.save(file)
    workbook.close()

    return file

def stromliste(intervall, df):
    # Erstelle Farhzeug Liste
    strom = pd.DataFrame({
        'Fg':[np.nan]*intervall,
        'Rad':[np.nan]*intervall,
        'K-Rad': [np.nan]*intervall,
        'Pkw':[np.nan]*intervall,
        'Lfw':[np.nan]*intervall,
        'Lkw':[np.nan]*intervall,
        'Bus':[np.nan]*intervall,
        'Lz/ Sz':[np.nan]*intervall
        }, index=df.index)
    
    return strom

def firstpage(file, minutes=15):
    """
    # --- Basistabellen erzeugen ---
    """
    intervall = int((60/minutes)*24)
    # DataFrame für die Fahrzeugtypen-Faktoren
    vehicle_factors = pd.DataFrame({
        'Faktor' : ['PKW-E', 'Kfz', 'SV'],
        'Fg': [np.nan, np.nan, np.nan],
        'Rad': [0.5, 1, np.nan],
        'K-Rad': [1, 1, np.nan],
        'Pkw': [1, 1, np.nan],
        'Lfw': [1, 1, np.nan],
        'Lkw': [1.5, 1, 1],
        'Bus': [1.5, 1, 1],
        'Lz/Sz': [2, np.nan, np.nan]
    })
         
    # Erstelle 15-Min-Zeitintervall von 00:00 bis 24:00
    start_time = pd.Timestamp("00:00")
    interval_15 = []
    for i in range(intervall):
        von = start_time + pd.Timedelta(minutes=15 * i)
        bis = von + pd.Timedelta(minutes=15)
        interval_15.append((von.strftime("%H:%M"), bis.strftime("%H:%M")))
    
    # 
    df_arm = pd.DataFrame({
        'PKW-E':[np.nan]*intervall,
        'Kfz': [np.nan]*intervall,
        'SV':[np.nan]*intervall
    })

    # DataFrame für die 15-Min-Intervalle
    df_15min = pd.DataFrame({
        ' ' : [np.nan]*intervall,
        'von': [i[0] for i in interval_15],
        'bis': [i[1] for i in interval_15],
        'PKW-E': [np.nan] * intervall,
        'Kfz': [np.nan] * intervall,
        'SV': [np.nan] * intervall,
        'SV(%)': [np.nan] * intervall
    })
    df_15min["VON_UHR"] = pd.to_datetime(df_15min["von"], format="%H:%M").dt.time # Uhrzeitenformat
    df_15min = df_15min.set_index("VON_UHR")

    """
    # --- Basisdaten verarbeiten ---
    """
    # Beschaffung der Werte
    df = pd.read_excel(file, sheet_name='DB-01', header=0)
    df["VON_UHR"] = pd.to_datetime(df["VON_UHR"], format="%H:%M:%S").dt.time # Uhrzeitenformat
    
    # Knotenarmanzahl // Grupieren
    arme = df.drop_duplicates(['VON_ARM','NACH_ARM'])[['VON_ARM','NACH_ARM']]
    grouped = df.groupby(['VON_ARM', 'NACH_ARM'])
    
    for _, row in tqdm(arme.iterrows()):
        von = row['VON_ARM']
        nach = row['NACH_ARM']
        # Zugriff auf eine Gruppe (z.B. A nach C)
        matrix = grouped[['VON_UHR', 'Fg', 'Rad', 'K-Rad', 'Pkw', 'Lfw', 'Lkw', 'Bus', 'Lz/Sz']].get_group((von, nach))
        # Hinzufügen der dynamischen Spalten
        matrix['PKW-E'] = np.nan
        matrix['KFZ'] = np.nan
        matrix['SV'] = np.nan
        # matrix = pd.concat([matrix,df_arm], axis=1) # Erste das Ende anbauen

        matrix["VON_UHR"] = pd.to_datetime(matrix["VON_UHR"], format="%H:%M:%S").dt.time # Uhrzeitenformat
        matrix = matrix.set_index("VON_UHR")
        matrix = matrix[~matrix.index.duplicated(keep='first')]
        # Vereinen der Dataframes
        df_15min = pd.concat([df_15min, matrix],axis=1,) # Dann das gesammte df anhängen
    
    """
    # --- Sheet erstellen ---
    """
    # Excel Sheet zusammen stellten
    workbook = load_workbook(file)
    workbook.create_sheet("01")
    worksheet = workbook['01']
    worksheet.title = "Verkehrszählung"

    # Infozeile erstellen & hinzufügen
    infos = {
       "Titel": "",
       "Ort": "",
       "Nummer": str(df.iloc[0]["KNOTENPUNKT"]),
       "Datum": df.iloc[0]["DATUM"]
       }

    row = 1
    for key, value in infos.items():
        worksheet[f'A{row}'] = f"{key}: {value}"
        row += 1
    # Leerzeile1 einfügen
    worksheet[f'A{row}'] = " "
    row += 1

    # Füge die Tabelle hinzu (ab Zeile 6)
    for idx, r in enumerate(dataframe_to_rows(vehicle_factors, index=False, header=True), start=1):
        worksheet.append(r)
    row += idx
    # Erfassen der Faktorzellen
    col_sv = row - 1
    col_kfz = row - 2
    col_pkw = row - 3
    # Leerzeile2 einfügen
    worksheet[f'A{row}'] = " " 
    row += 1
    worksheet[f'A{row}'] = " " 
    row += 1

    # DataFrame in Worksheet schreiben
    for r in dataframe_to_rows(df_15min, index=False, header=True):
        worksheet.append(r)

    """
    # --- Dynamische Formeln hinzufügen ---
    """
    # Annahme: Daten beginnen in Zeile 2 (weil Zeile 1 Header)
    start_row = row+1
    end_row = worksheet.max_row
    start_col = 4
    max_col = 11*len(arme) + 7
    col_ges_pkwe = get_column_letter(start_col)
    col_ges_kfz = get_column_letter(start_col+1)
    col_ges_sv = get_column_letter(start_col+2)
    col_ges_svpro = get_column_letter(start_col+3)

    # Formel "Summe"
    for row in tqdm(range(start_row, end_row + 1),leave=True, position=0):
        # Erstellung der Formlen in den ersten drei Spalten
        f_sv = f_kfz = f_pkwe = "=0"
        for arm in tqdm(range(len(arme)),leave=False, position=1):
            col = 16+11*arm
            f_pkwe += f"+{get_column_letter(col)}{row}"
            f_kfz += f"+{get_column_letter(col+1)}{row}"
            f_sv += f"+{get_column_letter(col+2)}{row}"
        # Einpflegen der Formeln
        worksheet[f"{col_ges_pkwe}{row}"] = f_pkwe # f"=R{row}+AC{row}+AN{row}+AY{row}+BJ{row}+BU{row}+CF{row}+CQ{row}+DB{row}+DM{row}+DX{row}+EI{row}+ET{row}+FE{row}+FP{row}+GA{row}"
        worksheet[f"{col_ges_kfz}{row}"] = f_kfz # f"=S{row}+AD{row}+AO{row}+AZ{row}+BK{row}+BV{row}+CG{row}+CR{row}+DC{row}+DN{row}+DY{row}+EJ{row}+EU{row}+FF{row}+FQ{row}+GB{row}"
        worksheet[f"{col_ges_sv}{row}"] = f_sv # f"=T{row}+AE{row}+AP{row}+BA{row}+BL{row}+BW{row}+CH{row}+CS{row}+DD{row}+DO{row}+DZ{row}+EK{row}+EV{row}+FG{row}+FR{row}+GC{row}"
        # SV% 
        # worksheet[f"{col_ges_svpro}{row}"].value = f"=IF({col_ges_kfz}{row}>0;{col_ges_sv}{row}/{col_ges_kfz}{row};0)"
        formula = f"=IF({col_ges_sv}{row}<>0,{col_ges_sv}{row}/{col_ges_kfz}{row},0)"
        worksheet[f"{col_ges_svpro}{row}"].value = formula

        # Erstellung der Formeln für die Arme
        for arm in tqdm(range(len(arme)),leave=False, position=1):
            # Basiswerte 
            col = 8 + 11*arm
            pos = {
                1 : get_column_letter(col), # letterFg
                2 : get_column_letter(col+1), # letterRad
                3 : get_column_letter(col+2), # letterKRad
                4 : get_column_letter(col+3), # letterPkw
                5 : get_column_letter(col+4), # letterLfw
                6 : get_column_letter(col+5), # letterLkw
                7 : get_column_letter(col+6), # letterBus
                8 : get_column_letter(col+7), # letterLz
                "sum_pkwe" : get_column_letter(col+8), 
                "sum_kfz" : get_column_letter(col+9), 
                "sum_sv" : get_column_letter(col+10)
            }
            f_arm_sv = f_arm_kfz = f_arm_pkwe = "=0"
            for i in tqdm(range(1,9),leave=False, position=1):
                f_arm_pkwe += f"+{pos[i]}{row}*${get_column_letter(i+1)}${col_pkw}"
                f_arm_kfz += f"+{pos[i]}{row}*${get_column_letter(i+1)}${col_kfz}"
                f_arm_sv += f"+{pos[i]}{row}*${get_column_letter(i+1)}${col_sv}"
            # Pkw-E
            worksheet[f"{pos['sum_pkwe']}{row}"].value = f_arm_pkwe 
            # f"{letterFg}{row}*$B${col_pkw}+{letterRad}{row}*$C${col_pkw}+{letterKRad}{row}*$D${col_pkw}+{letterPkw}{row}*$E${col_pkw}{letterLfw}{row}*$F${col_pkw}{letterLkw}{row}*$G${col_pkw}{letterBus}{row}*$H${col_pkw}{letterLz}{row}*$I${col_pkw}"
            # Kfz
            worksheet[f"{pos['sum_kfz']}{row}"].value = f_arm_kfz 
            # f"{letterFg}{row}*$B${col_kfz}+{letterRad}{row}*$C${col_kfz}+{letterKRad}{row}*$D${col_kfz}+{letterPkw}{row}*$E${col_kfz}{letterLfw}{row}*$F${col_kfz}{letterLkw}{row}*$G${col_kfz}{letterBus}{row}*$H${col_kfz}{letterLz}{row}*$I${col_kfz}"
            # Sv
            worksheet[f"{pos['sum_sv']}{row}"].value = f_arm_sv 
            # f"{letterFg}{row}*$B${col_sv}+{letterRad}{row}*$C${col_sv}+{letterKRad}{row}*$D${col_sv}+{letterPkw}{row}*$E${col_sv}{letterLfw}{row}*$F${col_sv}{letterLkw}{row}*$G${col_sv}{letterBus}{row}*$H${col_sv}{letterLz}{row}*$I${col_sv}"
    
    # Fußzeile einfügen
    worksheet.merge_cells(start_row=end_row+1, start_column=2, end_row=end_row+1, end_column=3)
    cell = worksheet[f"{get_column_letter(2)}{end_row+1}"]
    cell.value = "Gesammtsumme"
    cell.font = Font(bold=True)
    for col_idx in tqdm(range(4,max_col+1), leave=False, position=0):
        cell = worksheet[f"{get_column_letter(col_idx)}{end_row+1}"]
        cell.value = f"=SUM({get_column_letter(col_idx)}{start_row+1}:{get_column_letter(col_idx)}{end_row})"
        cell.font = Font(bold=True)
    worksheet[f"G{end_row+1}"].value = f"=F{end_row+1}/E{end_row+1}"

    """
    # --- Design Bereich ---
    """
    # alle Null-Werte ausblenden:
    # Benutzerdefiniertes Format definieren: Zeigt nichts bei 0 (0;-0;;@)
    # - 0: Positiv | - -0: Negativ | - ;;@: Für 0 und Text nichts anzeigen
    no_zero_style = NamedStyle(name='no_zero', number_format='0;-0;;@')
    # Format auf ALLE Zellen des Blattes anwenden (von Anfang an)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.style = no_zero_style

    # Prozentformat
    for cell in tqdm(worksheet["G"], leave=False, position=0):
        if cell.row >= start_row:  # Header(s) auslassen
            cell.number_format = "0.0%"
    
    # farbige Makierung - blau #CCFFFF
    start_row = 6
    end_row = start_row + 3 #9
    for col_idx in tqdm(range(1, 10), leave=False, position=0):
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        for row in range(start_row, end_row + 1):
            worksheet[f'{col_letter}{row}'].fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')
    
    # Umramung
    med_side = Side(style='medium', color='000000')
    thin_side = Side(style='thin', color='D9D9D9')
    thin_balck = Side(style='thin', color='000000')
    for col_idx in tqdm(range(1, 10), leave=False, position=0):
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        for row in range(start_row, end_row + 1):
            cell = worksheet[f'{col_letter}{row}']
            
            # Bestimme, welche Seiten einen Rahmen bekommen
            left = med_side if col_idx == 1 else thin_side
            right = med_side if col_idx == 9 else thin_side
            top = med_side if row == start_row else thin_side
            bottom = med_side if row == end_row else thin_side
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
    
    # Ergenzung Überschriften
    start_row = end_row + 1 #10
    mittel_row = start_row + 1 #11
    end_row = start_row + 2 #12

    worksheet.merge_cells(start_row=start_row, start_column=2, end_row=mittel_row, end_column=3)
    cell = worksheet[f'{get_column_letter(2)}{start_row}']
    cell.value = "Zeitintervall"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=7)
    cell = worksheet[f'{get_column_letter(4)}{start_row}']
    cell.value = "Zählstelle"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells(start_row=mittel_row, start_column=4, end_row=mittel_row, end_column=7)
    cell = worksheet[f'{get_column_letter(4)}{mittel_row}']
    cell.value = "Summe"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Beschriftung der Arme
    for idx, arm in tqdm(arme.iterrows(), leave=False, position=0):
        col = 11*idx + 8
        
        worksheet.merge_cells(start_row=start_row, start_column=col, end_row=start_row, end_column=col+10)
        cell = worksheet[f'{get_column_letter(col)}{start_row}']
        cell.value = "von/nach"
        cell.alignment = Alignment(horizontal='center', vertical='center')

        worksheet.merge_cells(start_row=mittel_row, start_column=col, end_row=mittel_row, end_column=col+10)
        cell = worksheet[f'{get_column_letter(col)}{mittel_row}']
        cell.value = f'{arm["VON_ARM"]}/{arm["NACH_ARM"]}'
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Font verändern
    for row in tqdm(range(start_row, end_row+1), leave=False, position=0):
        for col_idx in range(1, len(arme)*11+7):
            cell = worksheet[f'{get_column_letter(col_idx)}{row}']
            cell.font = Font(bold=True)
    
    # Erstellung der Abschnitts-Trennstellen
    trennerfirst = [2, 4,]
    trenner = []
    trenner2 = trennerfirst
    trennerNo = [3, 5, 6, 7,]
    for i in range(len(arme)+1):
        trenner.append(11*i+8)
        trenner2.append(11*i+16)
        trennerNo.append(11*i+17)
        trennerNo.append(11*i+18)
    # Kopfzeile
    # Oben
    for col_idx in tqdm(range(2, max_col+1), leave=False, position=0):
        cell = worksheet[f'{get_column_letter(col_idx)}{start_row}']
        if col_idx in trenner or col_idx in trennerfirst:
            cell.border = Border(top=med_side, bottom=None, left=med_side, right=None)
        else:
            cell.border = Border(top=med_side, bottom=None, left=None, right=None)
    # Mitte
    for col_idx in tqdm(range(2, max_col+1), leave=False, position=0):
        cell = worksheet[f'{get_column_letter(col_idx)}{mittel_row}']
        if col_idx in trenner or col_idx in trennerfirst:
            cell.border = Border(top=None, bottom=thin_balck, left=med_side, right=None)
        else:
            cell.border = Border(top=None, bottom=thin_balck, left=None, right=None)
    # Unten
    for col_idx in tqdm(range(2, max_col+1), leave=False, position=0):
        cell = worksheet[f'{get_column_letter(col_idx)}{end_row}']
        if col_idx in trenner or col_idx in trennerfirst:
            cell.border = Border(top=None, bottom=med_side, left=med_side, right=None)
        elif col_idx in trenner2:
                cell.border = Border(top=None, bottom=med_side, left=med_side, right=None)
        else:
            cell.border = Border(top=None, bottom=med_side, left=None, right=None)

    # farbige Makierung - grün #CCFFCC
    start_row = end_row + 1 #13
    end_row = start_row + 43 #56
    for col_idx in tqdm(range(2, 11*len(arme)+8), leave=False, position=0):
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        for row in range(start_row, end_row + 1):
            cell = worksheet[f'{col_letter}{row}']
            cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            if col_idx in trenner: # Umramung
                cell.border = Border(top=None, bottom=None, left=med_side, right=None) 
            elif col_idx in trenner2:
                cell.border = Border(top=None, bottom=thin_side, left=med_side, right=None)
            elif col_idx in trennerNo:
                cell.border = Border(top=None, bottom=thin_side, left=None, right=None)
            else: 
                cell.border = Border(top=None, bottom=None, left=None, right=None) 

    # farbige Makierung - gelb #FFFF99
    start_row = end_row + 1 #57
    end_row = start_row + 15 #72
    for col_idx in tqdm(range(2, 11*len(arme)+8), leave=False, position=0):
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        for row in range(start_row, end_row + 1):
            cell = worksheet[f'{col_letter}{row}']
            cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            if col_idx in trenner: # Umramung
                cell.border = Border(top=None, bottom=None, left=med_side, right=None) 
            elif col_idx in trenner2:
                cell.border = Border(top=None, bottom=thin_side, left=med_side, right=None)
            elif col_idx in trennerNo:
                cell.border = Border(top=None, bottom=thin_side, left=None, right=None)
            else: 
                cell.border = Border(top=None, bottom=None, left=None, right=None) 

    # farbige Makierung - blau #CCFFFF
    start_row = end_row + 1 #73
    end_row = start_row + 35 #108
    for col_idx in tqdm(range(2, 11*len(arme)+8), leave=False, position=0):
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        for row in range(start_row, end_row + 1):
            cell = worksheet[f'{col_letter}{row}']
            cell.fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')
            if col_idx in trenner: # Umramung
                cell.border = Border(top=None, bottom=None, left=med_side, right=None) 
            elif col_idx in trenner2:
                cell.border = Border(top=None, bottom=thin_side, left=med_side, right=None)
            elif col_idx in trennerNo:
                cell.border = Border(top=None, bottom=thin_side, left=None, right=None)
            else: 
                cell.border = Border(top=None, bottom=None, left=None, right=None) 

    # farbige Makierung - blau #99CCFF
    start_row = end_row + 1
    end_row = start_row
    for col_idx in tqdm(range(2, 11*len(arme)+8), leave=False, position=0):
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        for row in range(start_row, end_row + 1):
            cell = worksheet[f'{col_letter}{row}']
            cell.fill = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
            if col_idx in trenner:
                cell.border = Border(top=med_side, bottom=med_side, left=med_side, right=None)
            else:
                cell.border = Border(top=med_side, bottom=med_side, left=None, right=None)
    
    # Einfrieren der ersten 12 Zeilen
    worksheet.freeze_panes = "A13"

    """
    # --- Speicherung ---
    """
    file = file.replace('.xlsx','_test.xlsx', 1)
    workbook.save(file)
    workbook.close()

    # df_15min.to_excel(file)
    # for id, row in tqdm(df.iterrows()):
    #     matching_rows = df_15min[(df_15min['von'] == row['VON_UHR']) & 
    #                        (df_15min['VON_ARM'] == row['VON_ARM']) & 
    #                        (df_15min['NACH_ARM'] == row['NACH_ARM'])]
    #     if not matching_rows.empty:
    #         idx = matching_rows.index[0] 

    """
    # Erstelle 60-Min-Zeitntervall von 00:00 bis 24:00
    interval_60 = []
    for i in range(24):
        von = start_time + pd.Timedelta(hours= i)
        if intervall <= 24 & i == 23: # Für den 24:00 -> 00:00 Wechsel
            bis = pd.Timestamp("00:00") + pd.Timedelta(days=1)
        else:
            bis = von + pd.Timedelta(hours=1)
        interval_60.append((von.strftime("%H:%M"), bis.strftime("%H:%M")))

    

    # DataFrame für die 60-Min-Intervalle
    df_hourly = pd.DataFrame({
        'von': [i[0] for i in interval_60],
        'bis': [i[1] for i in interval_60],
        'PKW-E': [np.nan] * 24,
        'Kfz': [np.nan] * 24,
        'SV': [np.nan] * 24,
        'SV(%)': [np.nan] * 24
    })


    # Ausgabe der DataFrames (zum Überprüfen)
    print("15-Minuten-Intervalle DataFrame:")
    print(df_15min.head(10))  # Zeige die ersten 10 Zeilen
    print("\nStündliche Intervalle DataFrame:")
    print(df_hourly.head(10))
    print("\nFahrzeugtypen-Faktoren DataFrame:")
    print(vehicle_factors)
    df = pd.read_excel("F:\Rohdaten Kiel 2025.xlsx",sheet_name="01")
    print(df)
    """

if __name__ == "__main__":
    # 
    fs = [r"D:\Erhebungen\2026-01 Bochum Langendreh\Cam1\2026-02-01_13-17-45.counts_15min.csv",
          ]
    # fs = [r"F:\knoten2_2025-10-29_11-14-55.counts_15min_reduziert.csv",
    #       r"F:\knoten2_06_2025-10-29_10-07-42.counts_15min_reduziert.csv",
    #       r"F:\knoten2_21_2025-10-29_15-16-12.counts_15min_reduziert.csv"]
    d = connect(fs)
    f = convert(d, fs[0])
    firstpage(f)